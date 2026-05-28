"""
Excel (.xlsx) exporter for LIPSA results.

Uses openpyxl for good formatting.
This is an optional dependency — the exporter will raise a clear error if openpyxl is not installed.
"""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import Any

from lipsa.exporters.base import BaseExporter, ExportResult
from lipsa.models.post import Post


class ExcelExporter(BaseExporter):
    """Export posts to a formatted .xlsx file."""

    format_name = "excel"
    file_extension = ".xlsx"

    def export(
        self,
        records: Iterable[Post],
        output_path: str | Path,
        job_metadata: dict[str, Any] | None = None,
    ) -> ExportResult:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError(
                "Excel export requires the 'openpyxl' package.\n"
                "Install it with: pip install openpyxl"
            ) from None

        output_path = Path(output_path).with_suffix(self.file_extension)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "LIPSA Results"

        # Write metadata header if provided
        row = 1
        if job_metadata:
            ws.cell(row=row, column=1, value="LIPSA Export Metadata").font = Font(
                bold=True, size=14
            )
            row += 1
            for key, value in job_metadata.items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=str(value))
                row += 1
            row += 1  # blank row

        # Prepare data
        records_list = list(records)
        if not records_list:
            ws.cell(row=row, column=1, value="No records to export.")
            wb.save(output_path)
            return ExportResult(output_path=output_path, format=self.format_name, record_count=0)

        # Headers
        headers = [
            "post_urn",
            "url",
            "text",
            "author_name",
            "author_headline",
            "author_profile_url",
            "author_company",
            "posted_at",
            "reactions_count",
            "comments_count",
            "reposts_count",
            "content_type",
            "is_repost",
            "hashtags",
            "mentions",
        ]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Data rows
        for post in records_list:
            data = self._prepare_record(post)
            for col, header in enumerate(headers, 1):
                value = data.get(header, "")
                if isinstance(value, (list, dict)):
                    value = str(value)
                ws.cell(row=row, column=col, value=value)
            row += 1

        # Auto-adjust column widths (basic)
        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)

        return ExportResult(
            output_path=output_path,
            format=self.format_name,
            record_count=len(records_list),
            metadata=job_metadata,
        )
